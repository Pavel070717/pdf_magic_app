"""
Tests for routes/aocr.py — AOCR form generation.
"""

import json
from pathlib import Path
from unittest.mock import MagicMock, patch


class TestGenerateAOCR:
    def test_no_json_body(self, app_client):
        resp = app_client.post(
            "/api/aocr/generate",
            data="not json",
            content_type="text/plain",
        )
        assert resp.status_code == 400

    def test_empty_body(self, app_client):
        resp = app_client.post(
            "/api/aocr/generate",
            data=json.dumps({}),
            content_type="application/json",
        )
        assert resp.status_code == 400

    def test_template_not_found(self, app_client, temp_dir):
        with patch("routes.aocr.TEMPLATE", temp_dir / "nonexistent.xlsx"):
            resp = app_client.post(
                "/api/aocr/generate",
                data=json.dumps({"act_number": "1"}),
                content_type="application/json",
            )
            assert resp.status_code == 500

    def test_valid_generation(self, app_client, temp_dir):
        import shutil

        src = Path(__file__).parent.parent / "templates" / "AOCR_template.xlsx"
        if not src.exists():
            import openpyxl

            wb = openpyxl.Workbook()
            wb.create_sheet("1")
            wb.create_sheet("2")
            src.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(src))

        with patch("routes.aocr._get_output_dir", return_value=temp_dir):
            resp = app_client.post(
                "/api/aocr/generate",
                data=json.dumps(
                    {
                        "act_number": "42",
                        "act_date": "15.03.2026",
                        "object_name": "Test Object",
                        "developer_name": "Dev LLC",
                    }
                ),
                content_type="application/json",
            )
            assert resp.status_code == 200
            data = resp.get_json()
            assert data["success"] is True
            assert "xlsx_path" in data


class TestHelperFunctions:
    def test_unique_path_no_conflict(self, temp_dir):
        from routes.aocr import _unique_path

        result = _unique_path(temp_dir, "file", ".xlsx")
        assert result == temp_dir / "file.xlsx"

    def test_unique_path_with_conflict(self, temp_dir):
        from routes.aocr import _unique_path

        (temp_dir / "file.xlsx").touch()
        result = _unique_path(temp_dir, "file", ".xlsx")
        assert result == temp_dir / "file_1.xlsx"

    def test_unique_path_multiple_conflicts(self, temp_dir):
        from routes.aocr import _unique_path

        (temp_dir / "file.xlsx").touch()
        (temp_dir / "file_1.xlsx").touch()
        (temp_dir / "file_2.xlsx").touch()
        result = _unique_path(temp_dir, "file", ".xlsx")
        assert result == temp_dir / "file_3.xlsx"

    def test_fill_sheet1(self):
        from routes.aocr import SHEET1_MAP, _fill_sheet1

        mock_ws = MagicMock()
        data = {"object_name": "Test Object", "developer_name": "Dev LLC"}
        _fill_sheet1(mock_ws, data)
        assert mock_ws.__setitem__.call_count == 2

    def test_fill_sheet1_empty_data(self):
        from routes.aocr import _fill_sheet1

        mock_ws = MagicMock()
        _fill_sheet1(mock_ws, {})
        assert mock_ws.__setitem__.call_count == 0

    def test_fill_sheet2(self):
        from routes.aocr import _fill_sheet2

        mock_ws = MagicMock()
        data = {"s2_work_name": "Work Name", "s2_copies": "3"}
        _fill_sheet2(mock_ws, data)
        assert mock_ws.__setitem__.call_count == 2

    def test_fill_sheet2_empty_data(self):
        from routes.aocr import _fill_sheet2

        mock_ws = MagicMock()
        _fill_sheet2(mock_ws, {})
        assert mock_ws.__setitem__.call_count == 0


class TestExportPDF:
    def test_no_win32com(self, temp_dir):
        from routes.aocr import _export_pdf

        xlsx_path = temp_dir / "test.xlsx"
        xlsx_path.touch()
        with patch.dict("sys.modules", {"win32com": None, "win32com.client": None}):
            result = _export_pdf(xlsx_path)
            assert result is None
