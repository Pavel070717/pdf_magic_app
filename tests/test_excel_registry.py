"""
Tests for utils/excel_registry.py — pure functions and workbook helpers.
"""

from pathlib import Path

from openpyxl import Workbook

from utils.excel_registry import (
    build_header,
    build_info_block,
    build_signatures,
    fill_rows,
    merge_set,
    parse_filename,
    safe_filename,
    set_cell,
)


# ─── safe_filename ──────────────────────────────────────────────


class TestSafeFilename:
    def test_normal_text(self):
        assert safe_filename("hello") == "hello"

    def test_replaces_forbidden_chars(self):
        result = safe_filename('test/\\:*?"<>|file')
        assert "/" not in result
        assert "\\" not in result
        assert ":" not in result
        assert "*" not in result
        assert "?" not in result
        assert '"' not in result
        assert "<" not in result
        assert ">" not in result
        assert "|" not in result

    def test_preserves_valid_chars(self):
        assert safe_filename("file-name_v2.0") == "file-name_v2.0"

    def test_cyrillic_preserved(self):
        assert safe_filename("Документ") == "Документ"

    def test_empty_string(self):
        assert safe_filename("") == ""

    def test_numbers(self):
        assert safe_filename("12345") == "12345"


# ─── parse_filename ─────────────────────────────────────────────


class TestParseFilename:
    def test_standard_filename_with_semicolons(self):
        result = parse_filename("Акт;№1;05.02.2024.pdf")
        assert result["name"] != ""
        assert result["number"] == "№1"
        assert result["date"] == "05.02.2024"

    def test_filename_without_semicolons(self):
        result = parse_filename("Протокол испытаний.pdf")
        assert result["number"] == "-"
        assert result["date"] == "-"

    def test_filename_without_semicolons_with_date(self):
        result = parse_filename("Протокол 12.03.2025.pdf")
        assert result["date"] == "12.03.2025"

    def test_filename_with_prefix_number(self):
        result = parse_filename("01.АОСР;№1;05.02.2024.pdf")
        assert result["number"] == "№1"
        assert "АОСР" in result["name"]

    def test_filename_with_range_date(self):
        result = parse_filename("Документ;№2;01.01.2024-15.01.2024.pdf")
        assert result["date"] == "01.01.2024"
        assert result.get("date_end") == "15.01.2024"

    def test_filename_with_empty_number(self):
        result = parse_filename("Документ;;05.02.2024.pdf")
        assert result["number"] == "-"

    def test_filename_with_description(self):
        result = parse_filename("Документ;№1;05.02.2024;описание.pdf")
        assert result.get("description") == "описание"

    def test_filename_has_pages_and_page_num(self):
        result = parse_filename("Протокол испытаний.pdf")
        # No-semicolon branch always includes pages and page_num
        assert "pages" in result
        assert "page_num" in result

    def test_filename_preserves_original(self):
        result = parse_filename("Документ;№1;05.02.2024.pdf")
        assert result["filename"] == "Документ;№1;05.02.2024.pdf"

    def test_complex_cyrillic_name(self):
        result = parse_filename("01.Сертификат соответствия;Б-290;24.01.2026.pdf")
        assert "Сертификат" in result["name"]
        assert result["number"] == "Б-290"

    def test_filename_no_date_in_stem(self):
        result = parse_filename("Протокол испытаний.pdf")
        assert result["date"] == "-"

    def test_filename_date_with_dashes(self):
        result = parse_filename("Doc;1;01-01-2024.pdf")
        assert result["date"] == "01"


# ─── set_cell ────────────────────────────────────────────────────


class TestSetCell:
    def _ws(self):
        wb = Workbook()
        return wb.active

    def test_sets_value(self):
        ws = self._ws()
        set_cell(ws, "A1", "Hello")
        assert ws["A1"].value == "Hello"

    def test_sets_font(self):
        ws = self._ws()
        from utils.excel_registry import FONT14B
        set_cell(ws, "A1", "Bold", FONT14B)
        assert ws["A1"].font.bold is True

    def test_sets_alignment(self):
        ws = self._ws()
        from utils.excel_registry import ALIGN_L
        set_cell(ws, "A1", "Left", align=ALIGN_L)
        assert ws["A1"].alignment.horizontal == "left"

    def test_sets_border(self):
        ws = self._ws()
        from utils.excel_registry import THIN_BORDER
        set_cell(ws, "A1", "Border", border=THIN_BORDER)
        assert ws["A1"].border.left.style == "thin"

    def test_sets_fill(self):
        ws = self._ws()
        from utils.excel_registry import AQUA_FILL
        set_cell(ws, "A1", "Fill", fill=AQUA_FILL)
        assert ws["A1"].fill.start_color.rgb == "00C2F8FC"

    def test_no_border_or_fill(self):
        ws = self._ws()
        set_cell(ws, "A1", "Clean")
        assert ws["A1"].value == "Clean"


# ─── merge_set ──────────────────────────────────────────────────


class TestMergeSet:
    def _ws(self):
        wb = Workbook()
        return wb.active

    def test_merges_and_sets_value(self):
        ws = self._ws()
        merge_set(ws, "A1:C1", "Merged")
        assert ws["A1"].value == "Merged"

    def test_merged_cells_range(self):
        ws = self._ws()
        merge_set(ws, "A1:C1", "Test")
        assert "A1:C1" in [str(m) for m in ws.merged_cells.ranges]

    def test_applies_font(self):
        ws = self._ws()
        from utils.excel_registry import FONT12B
        merge_set(ws, "A1:B1", "Bold", FONT12B)
        assert ws["A1"].font.bold is True


# ─── build_info_block ──────────────────────────────────────────


class TestBuildInfoBlock:
    def _ws(self):
        wb = Workbook()
        return wb.active

    def test_returns_registry_number(self):
        ws = self._ws()
        data = {"registry_number": "42"}
        result = build_info_block(ws, data)
        assert result == "42"

    def test_sets_org_name(self):
        ws = self._ws()
        data = {"org_name": "ООО Строй"}
        build_info_block(ws, data)
        assert ws["B2"].value == "ООО Строй"

    def test_sets_object_name(self):
        ws = self._ws()
        data = {"object_name": "Объект Тест"}
        build_info_block(ws, data)
        assert ws["B5"].value == "Объект Тест"

    def test_sets_customer(self):
        ws = self._ws()
        data = {"customer": "Заказчик АО"}
        build_info_block(ws, data)
        assert ws["B9"].value == "Заказчик АО"

    def test_sets_registry_number_cell(self):
        ws = self._ws()
        data = {"registry_number": "123"}
        build_info_block(ws, data)
        assert ws[f"D{24}"].value == "123"

    def test_empty_data(self):
        ws = self._ws()
        result = build_info_block(ws, {})
        assert result == ""


# ─── build_header ───────────────────────────────────────────────


class TestBuildHeader:
    def _ws(self):
        wb = Workbook()
        return wb.active

    def test_sets_all_headers(self):
        ws = self._ws()
        build_header(ws)
        expected = ["№ п/п", "Наименование документа", "№ чертежа, акта, разрешения, журнала и др."]
        for exp in expected:
            found = False
            for row in range(26, 29):
                for col in "ABCDEFGH":
                    if ws[f"{col}{row}"].value == exp:
                        found = True
                        break
                if found:
                    break
            assert found, f"Header '{exp}' not found"

    def test_subheader_numbers(self):
        ws = self._ws()
        build_header(ws)
        for i in range(1, 9):
            col_letter = chr(64 + i)
            assert ws[f"{col_letter}28"].value == str(i)


# ─── build_signatures ──────────────────────────────────────────


class TestBuildSignatures:
    def _ws(self):
        wb = Workbook()
        return wb.active

    def test_returns_start_row(self):
        ws = self._ws()
        data = {"signature_sdal": "Иванов И.И."}
        result = build_signatures(ws, data, 29)
        assert isinstance(result, int)
        assert result > 29

    def test_sets_signature_labels(self):
        ws = self._ws()
        data = {}
        start = build_signatures(ws, data, 29)
        assert ws[f"A{start}"].value == "Сдал:"
        assert ws[f"A{start + 4}"].value == "Проверил:"
        assert ws[f"A{start + 8}"].value == "Принял:"

    def test_sets_representative_names(self):
        ws = self._ws()
        data = {"signature_sdal": "Петров П.П."}
        start = build_signatures(ws, data, 29)
        assert ws[f"C{start + 1}"].value == "Петров П.П."


# ─── fill_rows ─────────────────────────────────────────────────


class TestFillRows:
    def _ws(self):
        wb = Workbook()
        return wb.active

    def test_returns_last_row(self):
        ws = self._ws()
        rows = [
            {"name": "Doc1", "number": "N1", "date": "01.01.2024", "pages": 2, "page_num": "1-2", "filename": "f1.pdf"},
        ]
        result = fill_rows(ws, rows, Path("/tmp"))
        assert result >= 29

    def test_empty_rows(self):
        ws = self._ws()
        result = fill_rows(ws, [], Path("/tmp"))
        assert result == 28  # ROW_TABLE_DATA_START - 1

    def test_sets_cell_values(self):
        ws = self._ws()
        rows = [
            {"name": "Акт", "number": "№1", "date": "05.02.2024", "pages": 1, "page_num": "1", "filename": "act.pdf"},
        ]
        fill_rows(ws, rows, Path("/tmp"))
        assert ws["B29"].value == "Акт"
        assert ws["C29"].value == "№1"
        assert ws["D29"].value == "05.02.2024"

    def test_aosr_highlight(self):
        ws = self._ws()
        rows = [
            {"name": "АОСР. Стены", "number": "N1", "date": "01.01", "pages": 1, "page_num": "1", "filename": "a.xlsx"},
        ]
        fill_rows(ws, rows, Path("/tmp"))
        assert ws["A29"].value.startswith("=COUNTIF")

    def test_multiple_rows(self):
        ws = self._ws()
        rows = [
            {"name": f"Doc{i}", "number": f"N{i}", "date": "01.01", "pages": 1, "page_num": str(i), "filename": f"f{i}.pdf"}
            for i in range(5)
        ]
        result = fill_rows(ws, rows, Path("/tmp"))
        assert result == 33  # 29 + 5 - 1
