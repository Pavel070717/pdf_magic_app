"""
Blueprint: /api/aocr/* — Генерация формы АОСР (Акт освидетельствования скрытых работ).
Использует шаблон templates/AOCR_template.xlsx, заполняет через openpyxl,
сохраняет .xlsx и (опционально) .pdf через win32com.
"""

import shutil
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any

from flask import Blueprint, jsonify, request

from routes.core import logger, sanitize_text

aocr_bp = Blueprint("aocr", __name__)

TEMPLATE = Path(__file__).parent.parent / "templates" / "AOCR_template.xlsx"

# ─── Ячейки листа "1" ────────────────────────────────────────────────────────
SHEET1_MAP: dict[str, str] = {
    "object_name": "A6",
    "developer_name": "D9",
    "developer_address": "A10",
    "builder_name": "A13",
    "builder_continued": "A15",
    "builder_continued2": "A16",
    "designer_name": "A19",
    "designer_address": "A21",
    "designer_continued": "A23",
    "act_number": "C29",
    "act_date": "V29",
    "rep_developer": "A33",
    "rep_builder": "A36",
    "rep_builder_control": "A40",
    "rep_designer": "A45",
    "rep_contractor": "A50",
    "rep_others": "A53",
    "rep_others_continued": "A54",
}

# ─── Ячейки листа «2» ──────────────────────────────────────────────────────
SHEET2_MAP: dict[str, str] = {
    "s2_work_name": "A5",
    "s2_project_docs": "A9",
    "s2_materials_used": "A13",
    "s2_documents_submitted": "A17",
    "s2_start_day": "N20",
    "s2_start_month": "P20",
    "s2_start_year": "U20",
    "s2_end_day": "N21",
    "s2_end_month": "P21",
    "s2_end_year": "U21",
    "s2_standards_l1": "L23",
    "s2_standards_l2": "A24",
    "s2_standards_l3": "A25",
    "s2_standards_l4": "A26",
    "s2_standards_l5": "A27",
    "s2_next_work": "A31",
    "s2_additional_info": "I34",
    "s2_copies": "F36",
    "s2_appendices": "A39",
    "s2_rep_developer": "A44",
    "s2_rep_builder": "A48",
    "s2_rep_builder_ctrl": "A53",
    "s2_rep_designer": "A59",
    "s2_rep_contractor": "A65",
    "s2_rep_others": "A69",
}


# ─── Вспомогательные функции ─────────────────────────────────────────────────


def _get_output_dir() -> Path:
    """Возвращает директорию для сохранения из current_directory в state.json."""
    from utils.state import load_state

    state = load_state()
    current = state.get("current_directory")
    if current and current.get("project_dir"):
        out = Path(current["project_dir"])
    else:
        # fallback: сегодняшняя дата на рабочем столе
        date_str = datetime.now().strftime("%d.%m.%Y")
        out = Path.home() / "Desktop" / date_str / "АОСР"
    out.mkdir(parents=True, exist_ok=True)
    return out


def _unique_path(directory: Path, stem: str, ext: str) -> Path:
    """Генерирует уникальный путь для файла: stem.ext, stem_1.ext, stem_2.ext ..."""
    candidate = directory / f"{stem}{ext}"
    if not candidate.exists():
        return candidate
    i = 1
    while True:
        candidate = directory / f"{stem}_{i}{ext}"
        if not candidate.exists():
            return candidate
        i += 1


def _fill_sheet1(ws, data: dict[str, Any]) -> None:
    """Заполняет лист «1» данными из JSON."""
    for field, cell in SHEET1_MAP.items():
        value = data.get(field)
        if value is not None:
            ws[cell] = sanitize_text(str(value))


def _fill_sheet2(ws, data: dict[str, Any]) -> None:
    """Заполняет лист «2»: название работы, даты, нормативы, представители."""
    for field, cell in SHEET2_MAP.items():
        value = data.get(field)
        if value is not None:
            ws[cell] = sanitize_text(str(value))


def _export_pdf(xlsx_path: Path) -> Path | None:
    """Пытается экспортировать .xlsx → .pdf через win32com (Excel).
    Возвращает путь к PDF или None при ошибке."""
    pdf_path = xlsx_path.with_suffix(".pdf")
    excel = None
    wb = None

    # Early check — bail before COM init if win32com is not available
    try:
        import win32com.client
    except ImportError:
        logger.warning("win32com не установлен — пропускаем генерацию PDF")
        return None

    try:
        import pythoncom

        pythoncom.CoInitialize()
    except ImportError:
        pass

    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(xlsx_path))
        wb.ExportAsFixedFormat(0, str(pdf_path))  # 0 = xlTypePDF
        logger.info(f"PDF экспортирован: {pdf_path}")
        return pdf_path
    except Exception:
        logger.warning(f"Ошибка экспорта PDF через Excel: {traceback.format_exc()}")
        return None
    finally:
        # Always clean up COM objects — Dispatch may succeed even when
        # a later operation fails, leaving a zombie Excel process.
        if wb is not None:
            try:
                wb.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        try:
            import pythoncom

            pythoncom.CoUninitialize()
        except Exception:
            pass


# ─── Эндпоинты ────────────────────────────────────────────────────────────────


@aocr_bp.route("/api/aocr/generate", methods=["POST"])
def generate_aocr():
    """Генерирует .xlsx (и .pdf если доступен Excel) на основе данных из JSON."""
    try:
        import openpyxl
    except ImportError:
        return jsonify({"success": False, "error": "openpyxl не установлен"}), 500

    data = request.get_json(silent=True)
    if not data:
        return (
            jsonify({"success": False, "error": "Тело запроса должно быть JSON"}),
            400,
        )

    # ── Проверка шаблона ─────────────────────────────────────────────────
    if not TEMPLATE.exists():
        logger.error(f"Шаблон не найден: {TEMPLATE}")
        return (
            jsonify({"success": False, "error": f"Шаблон не найден: {TEMPLATE}"}),
            500,
        )

    # ── Определение имени и пути ─────────────────────────────────────────
    act_number = sanitize_text(str(data.get("act_number", "")))
    act_date = sanitize_text(str(data.get("act_date", "")))
    act_date_clean = act_date.replace("«", "").replace("»", "").replace('"', "").strip()
    file_stem = (
        f"АОСР_{act_number}_{act_date_clean}"
        if act_number
        else f"АОСР_{datetime.now():%Y%m%d_%H%M%S}"
    )
    file_stem = "".join(c for c in file_stem if c.isalnum() or c in "._- ") or "АОСР"

    output_dir = _get_output_dir()
    logger.info(f"Генерация АОСР: stem={file_stem}, output_dir={output_dir}")

    # ── Копирование и заполнение шаблона ────────────────────────────────
    xlsx_path = _unique_path(output_dir, file_stem, ".xlsx")
    shutil.copy2(str(TEMPLATE), str(xlsx_path))
    logger.info(f"Шаблон скопирован в: {xlsx_path}")

    try:
        wb = openpyxl.load_workbook(str(xlsx_path))
    except Exception as e:
        logger.exception(f"Ошибка открытия {xlsx_path}")
        return jsonify({"success": False, "error": f"Ошибка открытия .xlsx: {e}"}), 500

    try:
        # Лист 1
        if "1" in wb.sheetnames:
            _fill_sheet1(wb["1"], data)
        else:
            logger.warning("Лист «1» не найден в шаблоне")

        # Лист 2
        if "2" in wb.sheetnames:
            _fill_sheet2(wb["2"], data)
        else:
            logger.warning("Лист «2» не найден в шаблоне")

        wb.save(str(xlsx_path))
        logger.info(f"XLSX сохранён: {xlsx_path}")
    except Exception as e:
        logger.exception("Ошибка заполнения шаблона")
        try:
            wb.close()
        except Exception:
            pass
        return jsonify({"success": False, "error": f"Ошибка заполнения: {e}"}), 500
    finally:
        try:
            wb.close()
        except Exception:
            pass

    # ── Экспорт в PDF ────────────────────────────────────────────────────
    pdf_path = _export_pdf(xlsx_path)

    return jsonify(
        {
            "success": True,
            "xlsx_path": str(xlsx_path),
            "pdf_path": str(pdf_path) if pdf_path else None,
            "pdf_generated": pdf_path is not None,
        }
    )
