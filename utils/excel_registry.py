"""
Модуль генерации Excel-реестра исполнительной документации.
Использует openpyxl для полного контроля над форматированием.
Исправленная полная версия.
"""

import logging
import os
import re
import shutil
import tempfile
import xml.etree.ElementTree as ElementTree  # изменено с ET на ElementTree
import zipfile
from datetime import datetime
from pathlib import Path

from natsort import natsorted
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

from utils.rules import apply_symbol_rules, normalize_name

logger = logging.getLogger(__name__)
# ==========================================================
# СТИЛИ
# ==========================================================
THIN = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOTTOM_BORDER = Border(bottom=THIN)
NO_BORDER = Border()
AQUA_FILL = PatternFill(fill_type="solid", start_color="C2F8FC", end_color="C2F8FC")
FONT10 = Font(name="Times New Roman", size=10)
FONT12 = Font(name="Times New Roman", size=12)
FONT12B = Font(name="Times New Roman", size=12, bold=True)
FONT14B = Font(name="Times New Roman", size=14, bold=True)
FONT12_BLUE = Font(name="Times New Roman", size=12, bold=True, color="00008B")
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CC = Alignment(horizontal="centerContinuous", vertical="center", wrap_text=True)
ALIGN_CN = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_LN = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_R = Alignment(horizontal="right", vertical="center", wrap_text=True)
ALLOWED_EXTS = (
    ".pdf",
    ".jpg",
    ".jpeg",
    ".png",
    ".doc",
    ".docx",
    ".xls",
    ".xlsx",
    ".dwg",
    ".dxf",
)

# ==========================================================
# КОНСТАНТЫ РАЗМЕТКИ
# ==========================================================
ROW_ORG_NAME = 2
ROW_ORG_SUBTITLE = 3
ROW_OBJECT_NAME = 5
ROW_OBJECT_SUBTITLE = 6
ROW_CUSTOMER = 9
ROW_CUSTOMER_SUBTITLE = 10
ROW_SK_REPRESENTATIVE = 13
ROW_SK_REPRESENTATIVE_SUBTITLE = 14
ROW_GENERAL_CONTRACTOR = 17
ROW_GC_SUBTITLE = 18
ROW_WORK_EXECUTOR = 21
ROW_WE_SUBTITLE = 22
ROW_REGISTRY_NUM = 24
ROW_REGISTRY_TITLE = 25
ROW_TABLE_NOTE = 26
ROW_TABLE_HEADER = 27
ROW_TABLE_SUBHEADER = 28
ROW_TABLE_DATA_START = 29
PRINT_SCALE = 67
PRINT_TITLE_ROWS = "27:28"
MARGIN_LEFT_CM = 3.0
MARGIN_RIGHT_CM = 1.5
MARGIN_TOP_CM = 2.0
MARGIN_BOTTOM_CM = 2.0
MARGIN_HEADER_CM = 0.3
MARGIN_FOOTER_CM = 0.3
CM_TO_INCHES = 1 / 2.54
SIGNATURES_GAP = 2
SIGNATURE_BLOCK_HEIGHT = 4


# ==========================================================
# ВСПОМОГАТЕЛЬНЫЕ
# ==========================================================
def set_cell(ws, ref, value="", font=FONT12, align=ALIGN_C, border=None, fill=None):
    c = ws[ref]
    c.value = value
    c.font = font
    c.alignment = align
    if border:
        c.border = border
    if fill:
        c.fill = fill


def merge_set(ws, rng, value="", font=FONT12, align=ALIGN_C, border=None, fill=None):
    ws.merge_cells(rng)
    first = rng.split(":")[0]
    set_cell(ws, first, value, font, align, border, fill)


def safe_filename(text: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", str(text))


# ==========================================================
# ШАПКА
# ==========================================================
def build_info_block(ws, registry_data):
    for col in ("B", "C", "D", "E", "F"):
        set_cell(
            ws,
            f"{col}2",
            registry_data.get("org_name", "") if col == "B" else "",
            FONT12,
            ALIGN_CC,
            BOTTOM_BORDER,
        )
    set_cell(
        ws,
        f"C{ROW_ORG_SUBTITLE}",
        "(наименование строительной организации)",
        FONT10,
        ALIGN_CN,
        NO_BORDER,
    )
    for col in ("B", "C", "D", "E", "F"):
        set_cell(
            ws,
            f"{col}5",
            registry_data.get("object_name", "") if col == "B" else "",
            FONT12,
            ALIGN_CC,
            BOTTOM_BORDER,
        )
    set_cell(
        ws,
        f"C{ROW_OBJECT_SUBTITLE}",
        "(наименование объекта, шифр проекта)",
        FONT10,
        ALIGN_CN,
        NO_BORDER,
    )

    headers = [
        (ROW_CUSTOMER - 1, "Застройщик или технический заказчик:"),
        (
            ROW_SK_REPRESENTATIVE - 1,
            "Представитель службы строительного контроля технического заказчика:",
        ),
        (ROW_GENERAL_CONTRACTOR - 1, "Генподрядчик:"),
        (ROW_WORK_EXECUTOR - 1, "Исполнитель работ:"),
    ]
    for row, text in headers:
        set_cell(ws, f"A{row}", text, FONT12B, ALIGN_LN)
    for col in ("B", "C", "D", "E", "F"):
        set_cell(
            ws,
            f"{col}9",
            registry_data.get("customer", "") if col == "B" else "",
            FONT12,
            ALIGN_CC,
            BOTTOM_BORDER,
        )
    set_cell(
        ws,
        f"C{ROW_CUSTOMER_SUBTITLE}",
        "(наименование организации)",
        FONT10,
        ALIGN_CN,
        NO_BORDER,
    )
    for col in ("B", "C", "D", "E", "F"):
        set_cell(
            ws,
            f"{col}13",
            registry_data.get("sk_representative", "") if col == "B" else "",
            FONT12,
            ALIGN_CC,
            BOTTOM_BORDER,
        )
    set_cell(ws, "C14", "(наименование организации)", FONT10, ALIGN_CN, NO_BORDER)
    for cell_ref in (f"B{ROW_SK_REPRESENTATIVE + 1}", f"B{ROW_SK_REPRESENTATIVE + 2}"):
        ws[cell_ref].border = NO_BORDER
    for col in ("B", "C", "D", "E", "F"):
        set_cell(
            ws,
            f"{col}17",
            registry_data.get("general_contractor", "") if col == "B" else "",
            FONT12,
            ALIGN_CC,
            BOTTOM_BORDER,
        )
    set_cell(
        ws,
        f"C{ROW_GC_SUBTITLE}",
        "(наименование организации)",
        FONT10,
        ALIGN_CN,
        NO_BORDER,
    )
    for cell_ref in (f"B{ROW_GC_SUBTITLE}",):
        ws[cell_ref].border = NO_BORDER
    for col in ("B", "C", "D", "E", "F"):
        set_cell(
            ws,
            f"{col}21",
            registry_data.get("work_executor", "") if col == "B" else "",
            FONT12,
            ALIGN_CC,
            BOTTOM_BORDER,
        )
    set_cell(
        ws,
        f"C{ROW_WE_SUBTITLE}",
        "(наименование организации)",
        FONT10,
        ALIGN_CN,
        NO_BORDER,
    )
    for cell_ref in (f"B{ROW_WE_SUBTITLE}",):
        ws[cell_ref].border = NO_BORDER

    num = registry_data.get("registry_number", "")
    set_cell(ws, f"B{ROW_REGISTRY_NUM}", "", FONT12B, ALIGN_R)
    set_cell(ws, f"C{ROW_REGISTRY_NUM}", "РЕЕСТР №", FONT12B, ALIGN_R)
    set_cell(ws, f"D{ROW_REGISTRY_NUM}", num, FONT12B, ALIGN_C)
    set_cell(ws, f"B{ROW_REGISTRY_TITLE}", "", FONT14B, ALIGN_R)
    set_cell(ws, f"C{ROW_REGISTRY_TITLE}", "исполнительной", FONT14B, ALIGN_R)
    set_cell(ws, f"D{ROW_REGISTRY_TITLE}", "документации.", FONT14B, ALIGN_L)
    return num


# ==========================================================
# ТАБЛИЦА
# ==========================================================
def build_header(ws):
    set_cell(
        ws,
        f"G{ROW_TABLE_NOTE}",
        "Данный столбец не входит в бумажный носитель",
        FONT12_BLUE,
        ALIGN_C,
        THIN_BORDER,
    )
    headers = {
        f"A{ROW_TABLE_HEADER}": "№ п/п",
        f"B{ROW_TABLE_HEADER}": "Наименование документа",
        f"C{ROW_TABLE_HEADER}": "№ чертежа, акта, разрешения, журнала и др.",
        f"D{ROW_TABLE_HEADER}": "Дата чертежа, акта, разрешения, журнала и др.",
        f"E{ROW_TABLE_HEADER}": "Количество страниц",
        f"F{ROW_TABLE_HEADER}": "Страница по списку",
        f"G{ROW_TABLE_HEADER}": "Гиперссылка",
        f"H{ROW_TABLE_HEADER}": "Комментарий",
    }
    for cell, txt in headers.items():
        col = cell[0]
        font = FONT12_BLUE if col in ("G", "H") else FONT12B
        set_cell(ws, cell, txt, font, ALIGN_C, THIN_BORDER, AQUA_FILL)
    for i in range(1, 9):
        col = get_column_letter(i)
        font = FONT12_BLUE if col in ("G", "H") else FONT12B
        set_cell(
            ws,
            f"{col}{ROW_TABLE_SUBHEADER}",
            str(i),
            font,
            ALIGN_C,
            THIN_BORDER,
            AQUA_FILL,
        )


# ==========================================================
# ПАРСИНГ ИМЁН ФАЙЛОВ
# ==========================================================
def parse_filename(filename: str) -> dict:
    stem = Path(filename).stem
    if ";" not in stem:
        stem_no_num = re.sub(r"^(?:\d+\.\s*)+", "", stem)
        name = normalize_name(stem_no_num)
        date = "-"
        m = re.search(r"(\d{2})[.\-](\d{2})[.\-](\d{4})", filename)
        if m:
            date = f"{m.group(1)}.{m.group(2)}.{m.group(3)}"
        return {
            "name": name,
            "number": "-",
            "date": date,
            "pages": None,
            "page_num": None,
            "filename": filename,
        }
    parts = stem.split(";")
    raw_name = parts[0].strip()
    name = normalize_name(re.sub(r"^(?:\d+\.\s*)+", "", raw_name).strip())
    name = apply_symbol_rules(name)
    number = parts[1].strip() if len(parts) > 1 else "-"
    if not number:
        number = "-"
    number = apply_symbol_rules(number)
    date = "-"
    date_end = None
    if len(parts) > 2 and parts[2].strip():
        raw_date = parts[2].strip()
        if "-" in raw_date:
            date_parts = raw_date.split("-", 1)
            date = date_parts[0].strip()
            date_end = date_parts[1].strip()
        else:
            date = raw_date
    description = parts[3].strip() if len(parts) > 3 else ""
    return {
        "name": name,
        "number": number,
        "date": date,
        "date_end": date_end,
        "description": description,
        "filename": filename,
    }


# ==========================================================
# ЧТЕНИЕ A5 ИЗ XLSX (для АОСР)
# ==========================================================
def read_aosr_cell_from_xlsx(xlsx_path: Path) -> str:
    wb = None
    try:
        wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
        if len(wb.sheetnames) >= 1:
            ws = wb[wb.sheetnames[0]]
            val = ws["A81"].value
            if val is not None:
                return str(val).strip()
    except Exception as e:
        logger.warning(f"Ошибка чтения A81 из {xlsx_path}: {e}")
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
    return ""


# ==========================================================
# СБОР ФАЙЛОВ
# ==========================================================
def collect_files(folder_path: Path, filter_names: set | None = None):
    files = natsorted(
        [
            f
            for f in os.listdir(folder_path)
            if f.lower().endswith(ALLOWED_EXTS) and not f.startswith("00.Реестр")
        ]
    )
    if filter_names is not None:
        files = [f for f in files if f in filter_names]

    rows = []
    current_page = 1

    for f in files:
        path = folder_path / f
        ext = path.suffix.lower()
        pages = None
        if ext == ".pdf":
            try:
                reader = PdfReader(str(path))
                pages = len(reader.pages)
            except Exception as e:
                logger.warning(f"Ошибка чтения PDF {f}: {e}")
        if pages is not None:
            if pages == 1:
                page_num = str(current_page)
            else:
                page_num = f"{current_page}-{current_page + pages - 1}"
            current_page += pages
        else:
            page_num = str(current_page)
            current_page += 1

        parsed = parse_filename(f)
        name = parsed["name"]
        if name.upper().startswith("АОСР") and ext == ".xlsx":
            cell_text = read_aosr_cell_from_xlsx(path)
            if cell_text:
                name = f"АОСР. {cell_text}"
                logger.info(f"АОСР: заменили наименование на: {name}")
        rows.append(
            {
                "name": name,
                "number": parsed["number"],
                "date": parsed["date"],
                "pages": pages if pages else "—",
                "page_num": page_num,
                "filename": f,
            }
        )
    return rows


# ==========================================================
# ЗАПОЛНЕНИЕ
# ==========================================================
def fill_rows(ws, rows, folder_path):
    start = ROW_TABLE_DATA_START
    row_offset = ROW_TABLE_DATA_START - 1
    for i, item in enumerate(rows):
        row = start + i
        highlight = item["name"].upper().startswith("АОСР")
        font = FONT12B if highlight else FONT12
        fill = AQUA_FILL if highlight else None
        page_formula = (
            f"=IF(SUM($E$29:E{row})-E{row}+1=SUM($E$29:E{row}),"
            f"SUM($E$29:E{row}),"
            f'SUM($E$29:E{row})-E{row}+1&"-"&SUM($E$29:E{row}))'
        )
        if highlight:
            num_value = f'=COUNTIF(B${start}:B{row},"АОСР*")'
        else:
            num_value = f"=ROW()-{row_offset}"
        vals = {
            "A": num_value,
            "B": item["name"],
            "C": item["number"],
            "D": item["date"],
            "E": item["pages"],
            "F": page_formula,
            "G": item["filename"],
            "H": "",
        }
        for col, val in vals.items():
            align = ALIGN_L if col == "B" else ALIGN_C
            set_cell(ws, f"{col}{row}", val, font, align, THIN_BORDER, fill)
        cell = ws[f"G{row}"]
        cell.hyperlink = str(folder_path / item["filename"])
        cell.font = Font(
            name="Times New Roman", size=12, color="8E487F", italic=True, bold=highlight
        )
        cell.alignment = ALIGN_L
    return start + len(rows) - 1


# ==========================================================
# ПОДПИСИ
# ==========================================================
def build_signatures(ws, registry_data, last_row):
    start = last_row + SIGNATURES_GAP
    data = [
        (start, "Сдал:", "signature_sdal", "Подрядчика"),
        (
            start + SIGNATURE_BLOCK_HEIGHT,
            "Проверил:",
            "signature_proveril",
            "Технадзора",
        ),
        (
            start + SIGNATURE_BLOCK_HEIGHT * 2,
            "Принял:",
            "signature_prinyal",
            "Заказчика",
        ),
    ]
    for row, title, key, org in data:
        set_cell(ws, f"A{row}", title, FONT12B, ALIGN_LN)
        set_cell(ws, f"A{row + 1}", "Представитель", FONT12, ALIGN_LN)
        set_cell(ws, f"A{row + 2}", org, FONT12, ALIGN_LN)
        set_cell(
            ws,
            f"C{row + 1}",
            registry_data.get(key, ""),
            FONT12,
            ALIGN_C,
            BOTTOM_BORDER,
        )
        set_cell(ws, f"E{row + 2}", "(подпись)", FONT12, ALIGN_C, BOTTOM_BORDER)
        set_cell(ws, f"F{row + 2}", "(дата)", FONT12, ALIGN_C, BOTTOM_BORDER)
    top_border = Border(top=THIN)
    for block_row in (
        start + 2,
        start + SIGNATURE_BLOCK_HEIGHT + 2,
        start + SIGNATURE_BLOCK_HEIGHT * 2 + 2,
    ):
        set_cell(ws, f"E{block_row}", "(подпись)", FONT10, ALIGN_C, top_border)
        set_cell(ws, f"F{block_row}", "(дата)", FONT10, ALIGN_C, top_border)
    for block_row in (
        start + 2,
        start + SIGNATURE_BLOCK_HEIGHT + 2,
        start + SIGNATURE_BLOCK_HEIGHT * 2 + 2,
    ):
        set_cell(
            ws,
            f"C{block_row}",
            "(должность, фамилия, инициалы) М. П.",
            FONT10,
            ALIGN_C,
            NO_BORDER,
        )
    return start


# ==========================================================
# ПЕЧАТЬ
# ==========================================================
def setup_print(ws, end_row):
    ws.page_margins.left = MARGIN_LEFT_CM * CM_TO_INCHES
    ws.page_margins.right = MARGIN_RIGHT_CM * CM_TO_INCHES
    ws.page_margins.top = MARGIN_TOP_CM * CM_TO_INCHES
    ws.page_margins.bottom = MARGIN_BOTTOM_CM * CM_TO_INCHES
    ws.page_margins.header = MARGIN_HEADER_CM * CM_TO_INCHES
    ws.page_margins.footer = MARGIN_FOOTER_CM * CM_TO_INCHES
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.scale = PRINT_SCALE
    ws.print_title_rows = PRINT_TITLE_ROWS
    ws.print_area = f"A1:F{end_row + 10}"


def fix_page_margins_in_xlsx(xlsx_path):
    margins = {
        "left": str(MARGIN_LEFT_CM * CM_TO_INCHES),
        "right": str(MARGIN_RIGHT_CM * CM_TO_INCHES),
        "top": str(MARGIN_TOP_CM * CM_TO_INCHES),
        "bottom": str(MARGIN_BOTTOM_CM * CM_TO_INCHES),
    }
    temp_dir = None
    try:
        temp_dir = tempfile.mkdtemp()
        with zipfile.ZipFile(xlsx_path, "r") as zip_ref:
            zip_ref.extractall(temp_dir)
        worksheet_path = Path(temp_dir) / "xl" / "worksheets" / "sheet1.xml"
        if worksheet_path.exists():
            tree = ElementTree.parse(worksheet_path)  # заменено ET на ElementTree
            root = tree.getroot()
            for elem in root.iter(
                "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}pageMargins"
            ):
                elem.set("left", margins["left"])
                elem.set("right", margins["right"])
                elem.set("top", margins["top"])
                elem.set("bottom", margins["bottom"])
                elem.set("header", str(MARGIN_HEADER_CM * CM_TO_INCHES))
                elem.set("footer", str(MARGIN_FOOTER_CM * CM_TO_INCHES))
                logger.info(f"Обновлены поля страницы: {elem.attrib}")
            for elem in root.iter(
                "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}pageSetup"
            ):
                elem.set("scale", str(PRINT_SCALE))
                logger.info("Установлен масштаб 67%")
                for attr in ["fitToPage", "fitToHeight", "fitToWidth"]:
                    if attr in elem.attrib:
                        del elem.attrib[attr]
            tree.write(worksheet_path, encoding="utf-8", xml_declaration=True)
        fd, tmp_xlsx = tempfile.mkstemp(
            suffix=".xlsx", prefix="registry_", dir=str(xlsx_path.parent)
        )
        os.close(fd)
        try:
            with zipfile.ZipFile(tmp_xlsx, "w", zipfile.ZIP_DEFLATED) as zipf:
                for root_dir, dirs, files in os.walk(temp_dir):
                    for file in files:
                        file_path = Path(root_dir) / file
                        arcname = file_path.relative_to(temp_dir)
                        zipf.write(file_path, arcname)
            Path(tmp_xlsx).replace(xlsx_path)
        except Exception:
            try:
                Path(tmp_xlsx).unlink(missing_ok=True)
            except Exception:
                pass
            raise
        logger.info(f"Поля страницы успешно исправлены в {xlsx_path}")
    except Exception as e:
        logger.error(f"Ошибка при исправлении полей страницы: {e}")
        logger.exception("Детали ошибки:")
    finally:
        if temp_dir is not None:
            shutil.rmtree(temp_dir, ignore_errors=True)


# ==========================================================
# ОСНОВНАЯ ФУНКЦИЯ
# ==========================================================
def generate_excel_registry(
    folder_path: Path,
    registry_data: dict,
    saved_dicts: dict,
    filter_names: set | None = None,
) -> Path:
    logger.info(f"Генерация Excel реестра: {folder_path}")
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Реестр ВИ")
    ws.title = "Реестр ВИ"
    widths = {"A": 5, "B": 43, "C": 24, "D": 20, "E": 14, "F": 14, "G": 73, "H": 78}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    registry_number = build_info_block(ws, registry_data)
    build_header(ws)
    rows = collect_files(folder_path, filter_names)
    last_row = fill_rows(ws, rows, folder_path)
    sig_row = build_signatures(ws, registry_data, last_row)
    setup_print(ws, sig_row)
    today = datetime.now().strftime("%d%m-%Y")
    if registry_number:
        name = f"00.Реестр №{safe_filename(registry_number)}-{today}.xlsx"
    else:
        name = f"00.Реестр №{today}.xlsx"
    output = folder_path / name
    wb.save(output)
    logger.info(f"Реестр сохранён: {output}")
    try:
        fix_page_margins_in_xlsx(output)
    except Exception as e:
        logger.error(f"Ошибка при исправлении полей: {e}")
    return output
